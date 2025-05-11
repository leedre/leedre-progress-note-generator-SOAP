# Update:
# - Date entry color change from black to gray
# - removed pathlib
# - On program start, cursor will automatically be placed in the "Patient" entry field, ready for user input.
# - Cursor focus set to Patient entry field when the user clicks “New Patient” ready for the user to type Patient name.
# - Added extra space between the buttons and the bottom of the window.

import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import filedialog, simpledialog
from tkcalendar import Calendar
import random
from datetime import datetime
import pandas as pd
import json
import os
# from pathlib import Path
import sys
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Spacer
from tkinter import Menu

# Import sample data from sample_data.py
from sample_data import (
    chief_complaint, cpt_codes, units_options,
    subjective_samples, objective_samples, assessment_options
)

# Function to get the correct path for resources in both dev and bundled environments
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller."""
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller one-file mode: use temporary directory
        return os.path.join(sys._MEIPASS, relative_path)
    else:
        # Normal script execution: use script directory
        return os.path.join(os.path.abspath(os.path.dirname(__file__)), relative_path)

therapists = []  # Global list to store therapist names

# Password file handling
PASSWORD_FILE = resource_path("therapists_passwords.json")

# Global variable to store therapist passwords
therapist_passwords = {}

# Add global dictionary to store company information
company_info_dict = {}

# Global variable to store address options
address_options = []

# Add these to your existing global variables at the top
esigned_var = None  # Will be initialized as tk.StringVar later
esigners_by_company = {}  # Dictionary to store e-signers by company
esigned_date_entry = None    # New global variable for the e-signed date

# Update initialize_passwords to use the global therapists
def initialize_passwords():
    global therapist_passwords
    try:
        if os.path.exists(PASSWORD_FILE):
            with open(PASSWORD_FILE, "r") as file:
                therapist_passwords = json.load(file)
        else:
            # Default passwords if file doesn't exist, only if therapists list is populated
            if therapists:
                therapist_passwords = {therapist: "password" for therapist in therapists}
                if not hasattr(sys, '_MEIPASS'):
                    with open(PASSWORD_FILE, "w") as file:
                        json.dump(therapist_passwords, file)
            else:
                therapist_passwords = {}  # Empty dict if no therapists yet
    except Exception as e:
        # Fallback to empty dict or default passwords if therapists exist
        therapist_passwords = {therapist: "password" for therapist in therapists} if therapists else {}
        messagebox.showwarning("Warning", f"Could not load/save passwords: {e}. Using default passwords.")

# Function to save therapist passwords (only in dev mode)
def save_passwords():
    if not hasattr(sys, '_MEIPASS'):  # Only save in dev mode, not in bundle
        try:
            with open(PASSWORD_FILE, "w") as file:
                json.dump(therapist_passwords, file)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save passwords: {e}")

# Function to open the change password window
def open_change_password_window():
    def change_password():
        username = username_var.get()
        old_password = old_password_entry.get()
        new_password = new_password_entry.get()
        confirm_password = confirm_password_entry.get()

        if username in therapist_passwords and therapist_passwords[username] == old_password:
            if new_password == confirm_password:
                therapist_passwords[username] = new_password
                save_passwords()
                messagebox.showinfo("Password Changed", "Password changed successfully")
                change_password_window.destroy()
            else:
                messagebox.showerror("Error", "New passwords do not match")
        else:
            messagebox.showerror("Error", "Invalid username or old password")

    def return_to_login():
        change_password_window.destroy()
        login_window.deiconify()

    change_password_window = tk.Toplevel(root)
    change_password_window.title("Change Password")
    change_password_window.geometry("500x200")
    change_password_window.resizable(False, False)

    # Center the window
    window_width, window_height = 400, 300
    screen_width = change_password_window.winfo_screenwidth()
    screen_height = change_password_window.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    change_password_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    tk.Label(change_password_window, text="Username:").grid(row=0, column=0, padx=10, pady=10)
    username_var = tk.StringVar()
    username_dropdown = ttk.Combobox(change_password_window, textvariable=username_var, values=therapists, width=20)
    username_dropdown.grid(row=0, column=1, padx=10, pady=10)
    username_dropdown.current(0)

    tk.Label(change_password_window, text="Old Password:").grid(row=1, column=0, padx=10, pady=10)
    old_password_entry = tk.Entry(change_password_window, show="*", width=22)
    old_password_entry.grid(row=1, column=1, padx=10, pady=10)

    tk.Label(change_password_window, text="New Password:").grid(row=2, column=0, padx=10, pady=10)
    new_password_entry = tk.Entry(change_password_window, show="*", width=22)
    new_password_entry.grid(row=2, column=1, padx=10, pady=10)

    tk.Label(change_password_window, text="Confirm Password:").grid(row=3, column=0, padx=10, pady=10)
    confirm_password_entry = tk.Entry(change_password_window, show="*", width=22)
    confirm_password_entry.grid(row=3, column=1, padx=10, pady=10)

    tk.Button(change_password_window, text="Change Password", command=change_password).grid(row=4, column=0, columnspan=2, pady=10)
    tk.Button(change_password_window, text="Return to Login", command=return_to_login).grid(row=5, column=0, columnspan=2, pady=10)


# Function to load a new Excel file and update the chief complaint and username dropdowns
def load_excel_file():
    global current_file_path, df, workbook, therapist_passwords, therapist_license_dict, company_info_dict, therapists, address_options, esigners_by_company
    # Define the path to the Excel file in the Documents folder
    file_path = os.path.expanduser("~/PT note contents 20241004.xlsx")  # Replace 'data.xlsx' with your Excel file name
    
    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"Excel file not found at {file_path}. Please place this file in the home folder.")
        return

    try:
        # Read the Excel file to get sheet names
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        current_file_path = file_path

        # Load the PT_info sheet for therapist names, license numbers, company names, and e-signers
        if "PT_info" in sheet_names:
            pt_info_sheet = workbook["PT_info"]
            therapist_license_dict = {}
            therapists = []  # Reset and populate global therapists list
            company_names = []
            esigners_by_company = {}  # Reset e-signers dictionary

            for row in pt_info_sheet.iter_rows(min_row=2, values_only=True):
                company_name = row[0]  # "Company" in column A (index 0)
                esigner = row[1]  # E-signer in column B (index 1)
                treating_provider = row[2]  # "Treating provider" in column C (index 2)
                license_number = row[3]  # "License number" in column D (index 3)
                
                if company_name and company_name not in company_names:
                    company_names.append(company_name)
                
                # Populate esigners_by_company dictionary
                if company_name:
                    if company_name not in esigners_by_company:
                        esigners_by_company[company_name] = []
                    if esigner and esigner not in esigners_by_company[company_name]:
                        esigners_by_company[company_name].append(esigner)
                
                if treating_provider and treating_provider not in therapists:
                    therapists.append(treating_provider)
                    therapist_license_dict[treating_provider] = license_number

            username_dropdown['values'] = therapists
            if therapists:
                username_dropdown.current(0)

            for therapist in therapists:
                if therapist not in therapist_passwords:
                    therapist_passwords[therapist] = "password"
            save_passwords()

            therapist_dropdown_bottom['values'] = therapists
            if therapists:
                therapist_var.set(therapists[0])

            # Update the Company dropdown with company names from PT_info
            company_dropdown['values'] = company_names
            if company_names:
                company_dropdown.current(0)
                # Update E-signed by dropdown based on initial company selection
                initial_company = company_names[0]
                esigned_dropdown['values'] = esigners_by_company.get(initial_company, [])
                if esigners_by_company.get(initial_company):
                    esigned_var.set(esigners_by_company[initial_company][0])

            initialize_passwords()

        # Load the Address sheet for address dropdown and company_info_dict
        if "Address" in sheet_names:
            address_sheet = workbook["Address"]
            address_options = []  # Reset address options
            company_info_dict = {}  # Reset company_info_dict

            for row in address_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                address = row[0]  # Column A (index 0) - Address
                telephone = row[1]  # Column B (index 1) - Telephone
                fax = row[2]  # Column C (index 2) - Fax
                
                if address:
                    if address not in address_options:
                        address_options.append(address)
                    company_info_dict[address] = {
                        "address": address or "",
                        "telephone": telephone or "",
                        "fax": fax or ""
                    }
            
            address_dropdown['values'] = address_options
            if address_options:
                address_dropdown.current(0)

        # Load the Insurance sheet
        if "Insurance" in sheet_names:
            insurance_sheet = workbook["Insurance"]
            insurance_names = [cell.value for cell in insurance_sheet["A"][1:] if cell.value]
            insurance_dropdown['values'] = insurance_names
            if insurance_names:
                insurance_dropdown.current(0)

        # Update the chief complaint dropdown
        filtered_sheet_names = [sheet for sheet in sheet_names if sheet not in ["PT_info", "Address", "Insurance", "Low back for test"]]
        chief_complaint_dropdown['values'] = filtered_sheet_names or [""]
        chief_complaint_var.set(filtered_sheet_names[0] if filtered_sheet_names else "")

        load_data_based_on_chief_complaint()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while loading the Excel file: {e}")


# Login Screen
def login_screen():
    global login_window, username_dropdown
    login_window = tk.Toplevel(root)
    login_window.title("Login")
    login_window.geometry("400x300")
    login_window.resizable(False, False)

    # Center the login window
    window_width, window_height = 400, 300
    screen_width = login_window.winfo_screenwidth()
    screen_height = login_window.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    login_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    def validate_login():
        username = username_var.get()
        password = password_entry.get()

        if username in therapist_passwords and therapist_passwords[username] == password:
            login_window.destroy()
            root.deiconify()
            therapist_var.set(username)
            therapist_dropdown_bottom.config(state="enabled")  # Only bottom dropdown is disabled
            reset_all_fields()
            patient_entry.focus_set()  # Set focus to the patient_entry field
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")

# Add empty label for extra padding at the top
    tk.Label(login_window, text="").grid(row=0, column=0, columnspan=2, pady=1)  # Adjust pady for desired padding

    tk.Label(login_window, text="Username:").grid(row=1, column=0, padx=10, pady=10)
    username_var = tk.StringVar()
    username_dropdown = ttk.Combobox(login_window, textvariable=username_var, width=20)
    username_dropdown.grid(row=1, column=1, padx=10, pady=10)

    tk.Label(login_window, text="Password:").grid(row=2, column=0, padx=10, pady=10)
    password_entry = tk.Entry(login_window, show="*", width=22)
    password_entry.grid(row=2, column=1, padx=10, pady=10)

    tk.Button(login_window, text="Change Password", command=open_change_password_window).grid(row=4, column=1, padx=(10, 5), pady=10)
    tk.Button(login_window, text="Exit", command=exit_program).grid(row=4, column=0, padx=(5, 10), pady=10)

    password_entry.bind("<Return>", lambda event: validate_login())
    root.withdraw()

    # Automatically load the Excel file after widgets are created
    load_excel_file()

# Function to start dragging a row
def start_drag(event):
    global dragged_item
    # Identify the row under the mouse pointer
    dragged_item = cpt_tree.identify_row(event.y)
    if dragged_item:
        # Highlight the dragged row
        cpt_tree.selection_set(dragged_item)

# Global variable to store the workbook and sheet
workbook = None
sheet = None

# Global variable to store the login window
login_window = None

# Global variables to track drag-and-drop state
dragged_item = None

# Global variables to track drag-and-drop state for the procedure table
dragged_procedure_item = None

# Global variables to track drag-and-drop and visual line
visual_line = None

# Function to start dragging a row in the procedure table
def start_drag_procedure(event):
    global dragged_procedure_item
    # Identify the row under the mouse pointer
    dragged_procedure_item = procedure_table.identify_row(event.y)
    if dragged_procedure_item:
        # Highlight the dragged row
        procedure_table.selection_set(dragged_procedure_item)

# Function to handle dragging motion in the procedure table
def on_drag_procedure(event):
    global dragged_procedure_item
    if dragged_procedure_item:
        # Identify the row under the mouse pointer during the drag
        target_row = procedure_table.identify_row(event.y)
        if target_row:
            # Highlight the target row
            procedure_table.selection_set(target_row)

# Function to handle dropping a row in the procedure table
def stop_drag_procedure(event):
    global dragged_procedure_item
    if dragged_procedure_item:
        # Identify the target row where the dragged row is being dropped
        target_row = procedure_table.identify_row(event.y)
        if target_row and dragged_procedure_item != target_row:
            # Get the values of the dragged row
            dragged_values = procedure_table.item(dragged_procedure_item, "values")
            
            # Remove the dragged row from the table
            procedure_table.delete(dragged_procedure_item)
            
            # Insert the dragged row above the target row
            procedure_table.insert("", procedure_table.index(target_row), values=dragged_values)
            
            # Clear the dragged item
            dragged_procedure_item = None

# Function to start dragging a row
def start_drag(event):
    global dragged_item
    # Identify the row under the mouse pointer
    dragged_item = cpt_tree.identify_row(event.y)
    if dragged_item:
        # Highlight the dragged row
        cpt_tree.selection_set(dragged_item)

# Function to handle dragging motion
def on_drag(event):
    global dragged_item
    if dragged_item:
        # Identify the row under the mouse pointer during the drag
        target_row = cpt_tree.identify_row(event.y)
        if target_row:
            # Highlight the target row
            cpt_tree.selection_set(target_row)

# Function to handle dropping a row
def stop_drag(event):
    global dragged_item
    if dragged_item:
        # Identify the target row where the dragged row is being dropped
        target_row = cpt_tree.identify_row(event.y)
        if target_row and dragged_item != target_row:
            # Get the values of the dragged row
            dragged_values = cpt_tree.item(dragged_item, "values")
            
            # Remove the dragged row from the tree
            cpt_tree.delete(dragged_item)
            
            # Insert the dragged row above the target row
            cpt_tree.insert("", cpt_tree.index(target_row), values=dragged_values)
            
            # Clear the dragged item
            dragged_item = None

# Function to return to the login screen
def return_to_login():
    global login_window  # Declare login_window as global

    # Clear all entries and reset fields to default
    patient_entry.delete(0, tk.END)
    # dos_entry.delete(0, tk.END)
    dob_entry.delete(0, tk.END)
    # chief_complaint_dropdown.current(0 if chief_complaint_dropdown['values'] else -1)  # Set to 0 if values exist, else -1
    # if insurance_dropdown['values']:
        # insurance_dropdown.current(0)
    # chief_complaint_dropdown.current(0)
    cpt_var.set("")
    units_var.set("")
    units_dropdown.current(0)
    subjective_text.delete(1.0, tk.END)
    objective_text.delete(1.0, tk.END)
    # Function to clear all entries in the CPT code tree
    for row in cpt_tree.get_children():
        cpt_tree.delete(row)
    for row in procedure_table.get_children():
        procedure_table.delete(row)
    assessment_dropdown.current(2)
    plan_text.delete(1.0, tk.END)
    plan_text.insert(tk.END, "It is recommended that the patient continue physical therapy targeting the affected areas at a frequency of 2-3 sessions per week, with a follow-up re-evaluation in 4-5 weeks to assess progress and guide ongoing therapeutic management.")
    # therapist_date_entry.delete(0, tk.END)

    # Only set current if there are values in the dropdowns
    # if therapist_dropdown_top['values']:
    #     therapist_dropdown_top.current(0)
    if therapist_dropdown_bottom['values']:
        therapist_dropdown_bottom.current(0)

    root.withdraw()  # Hide the main application window
    login_screen()  # Show the login screen

# Function to exit the program
def exit_program():
    root.quit()

# Function to load the sheet based on the selected chief complaint
def load_sheet_based_on_chief_complaint(*args):
    global sheet, workbook
    selected_complaint = chief_complaint_var.get().strip()
    if not selected_complaint or not workbook:  # Skip if no complaint or no workbook loaded
        return
    if selected_complaint in workbook.sheetnames:  # Only load if it’s a valid sheet
        try:
            sheet = workbook[selected_complaint]
            load_data_based_on_chief_complaint()  # Optionally load data if needed
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet '{selected_complaint}': {e}")
    # If not a valid sheet, do nothing (allow manual entry to persist)


# def load_sheet_based_on_chief_complaint():
#     global sheet
#     selected_complaint = chief_complaint_var.get()
#     if selected_complaint:
#         try:
#             sheet = workbook[selected_complaint]  # Dynamically select the sheet based on the chief complaint
#         return
# New updated one
# def load_sheet_based_on_chief_complaint(*args):
#     global df
#     selected_sheet = chief_complaint_var.get().strip()
#     if not selected_sheet or not workbook:
#         return
#     if selected_sheet in workbook.sheetnames:
#         try:
#             df = pd.read_excel(current_file_path, sheet_name=selected_sheet)
#             update_checkboxes()  # Replace with your UI update function
#         except Exception as e:
#             print(f"Error loading sheet '{selected_sheet}': {e}")
    # Custom text is preserved in chief_complaint_var if no sheet exists

# Function to load data based on the selected chief complaint
def load_data_based_on_chief_complaint():
    global df
    selected_complaint = chief_complaint_var.get().strip()
    
    if not current_file_path or not selected_complaint:
        return  # Skip if no file loaded or no complaint selected

    if selected_complaint in workbook.sheetnames:  # Only load data if it’s a valid sheet
        try:
            df = pd.read_excel(current_file_path, sheet_name=selected_complaint)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    # If not a valid sheet, keep df unchanged or reset it if desired


# Function to generate random notes for Subjective
def generate_subjective():
    subjective_text.delete(1.0, tk.END)
    subjective_text.insert(tk.END, random.choice(subjective_samples))

# Function to generate random notes for Objective
def generate_objective():
    objective_text.delete(1.0, tk.END)
    objective_text.insert(tk.END, random.choice(objective_samples))

# Function to open calendar for DOB selection
def open_calendar():
    def set_date():
        dob_entry.delete(0, tk.END)
        dob_entry.insert(0, cal.selection_get().strftime("%m/%d/%Y"))
        top.destroy()

    top = tk.Toplevel(root)
    cal = Calendar(top, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(pady=10)
    ttk.Button(top, text="Select Date", command=set_date).pack(pady=10)

# Function to open calendar for DOS date selection
# Function to open calendar for DOS date selection and sync all dates
def open_dos_calendar():
    def set_date():
        selected_date = cal.selection_get().strftime("%m/%d/%Y")
        # Update DOS entry
        dos_entry.delete(0, tk.END)
        dos_entry.insert(0, selected_date)
        # Sync therapist date
        therapist_date_entry.delete(0, tk.END)
        therapist_date_entry.insert(0, selected_date)
        # Sync e-signed date
        esigned_date_entry.delete(0, tk.END)
        esigned_date_entry.insert(0, selected_date)
        top.destroy()

    top = tk.Toplevel(root)
    cal = Calendar(top, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(pady=10)
    ttk.Button(top, text="Select Date", command=set_date).pack(pady=10)

# Function to open calendar for therapist session date selection
def open_therapist_calendar():
    def set_date():
        therapist_date_entry.delete(0, tk.END)
        therapist_date_entry.insert(0, cal.selection_get().strftime("%m/%d/%Y"))
        top.destroy()

    top = tk.Toplevel(root)
    cal = Calendar(top, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(pady=10)
    ttk.Button(top, text="Select Date", command=set_date).pack(pady=10)

def open_esigned_calendar():
    def set_date():
        esigned_date_entry.delete(0, tk.END)
        esigned_date_entry.insert(0, cal.selection_get().strftime("%m/%d/%Y"))
        top.destroy()

    top = tk.Toplevel(root)
    cal = Calendar(top, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(pady=10)
    ttk.Button(top, text="Select Date", command=set_date).pack(pady=10)

# Function to set DOS to current date
def set_current_date():
    dos_entry.delete(0, tk.END)
    dos_entry.insert(0, datetime.now().strftime("%m/%d/%Y"))

# Function to clear the Procedure table
def clear_procedure():
    for row in procedure_table.get_children():
        procedure_table.delete(row)

# Function to clear the Subjective and Objective field
def clear_scenarios():
    subjective_text.delete(1.0, tk.END)
    objective_text.delete(1.0, tk.END)

# Function to clear the Subjective field
def clear_subjective():
    subjective_text.delete(1.0, tk.END)

# Function to clear the Objective field
def clear_objective():
    objective_text.delete(1.0, tk.END)

# Function to add a new procedure line
def add_procedure():
    new_procedure = procedure_entry.get()
    if new_procedure:
        procedure_table.insert("", "end", values=("", new_procedure))  # Empty CPT Code column
        procedure_entry.delete(0, tk.END)

def get_scenarios_from_excel():
    scenarios = {}
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):  # Skip header row
        # Scenario levels
        scenario_name = row[0]
        scenario_type = row[1]
        severity_9_10_subjective = row[2]
        severity_7_8_subjective = row[3]
        severity_5_6_subjective = row[4]
        severity_4_5_subjective = row[5]
        severity_2_3_subjective = row[6]

        if scenario_name not in scenarios:
            scenarios[scenario_name] = {
                "9-10": {"Subjective": "", "Objective": ""},
                "7-8": {"Subjective": "", "Objective": ""},
                "5-6": {"Subjective": "", "Objective": ""},
                "4-5": {"Subjective": "", "Objective": ""},
                "2-3": {"Subjective": "", "Objective": ""},
            }

        if scenario_type == "S":
            scenarios[scenario_name]["9-10"]["Subjective"] = severity_9_10_subjective
            scenarios[scenario_name]["7-8"]["Subjective"] = severity_7_8_subjective
            scenarios[scenario_name]["5-6"]["Subjective"] = severity_5_6_subjective
            scenarios[scenario_name]["4-5"]["Subjective"] = severity_4_5_subjective
            scenarios[scenario_name]["2-3"]["Subjective"] = severity_2_3_subjective
        elif scenario_type == "O":
            scenarios[scenario_name]["9-10"]["Objective"] = severity_9_10_subjective
            scenarios[scenario_name]["7-8"]["Objective"] = severity_7_8_subjective
            scenarios[scenario_name]["5-6"]["Objective"] = severity_5_6_subjective
            scenarios[scenario_name]["4-5"]["Objective"] = severity_4_5_subjective
            scenarios[scenario_name]["2-3"]["Objective"] = severity_2_3_subjective

    return scenarios

# Function to display a random scenario
def display_scenario():
    severity = severity_var.get()
    print(f"Debugging function display severity level: {severity}")
    if severity not in ["9-10", "7-8", "5-6", "4-5", "2-3"]:
        messagebox.showerror("Error", "Please select a severity level.")
        return

    # Load scenarios from the Excel file
    scenarios = get_scenarios_from_excel()

    # Randomly select a scenario
    scenario = random.choice(list(scenarios.keys()))
    print(f"Debugging function display scene no from:{scenario}")
    subjective = scenarios[scenario][severity]["Subjective"]
    objective = scenarios[scenario][severity]["Objective"]

    # Update the text fields
    subjective_text.delete(1.0, tk.END)
    subjective_text.insert(tk.END, subjective)
    objective_text.delete(1.0, tk.END)
    objective_text.insert(tk.END, objective)

def add_cpt_row():
    selected_cpt = cpt_var.get()
    selected_units = units_var.get()
    selected_complaint = chief_complaint_var.get()

    global df
    if current_file_path:
        try:
            df = pd.read_excel(current_file_path, sheet_name=selected_complaint)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet '{selected_complaint}': {e}")
            return

    if selected_complaint == "Neck (M54.2)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 14)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 8)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set(0)  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
################ R_Sh #####################
###########################################
    elif selected_complaint == "Rt Shoulder (M25.511)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 16)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 15)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
################ L_Sh #####################
###########################################
    elif selected_complaint == "Lt Shoulder (M25.512)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 16)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 15)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing
###########################################
################ R_El #####################
###########################################

    elif selected_complaint == "Rt Elbow (M25.521)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 12)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 13)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 9)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 11)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
################ L_El #####################
###########################################

    elif selected_complaint == "Lt Elbow (M25.522)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 12)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 13)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 9)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 11)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
################ R_W #####################
###########################################

    elif selected_complaint == "Rt Wrist (M25.531)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 7)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
################ L_W #####################
###########################################

    elif selected_complaint == "Lt Wrist (M25.532)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 7)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### R_Hand (M79.641) ########
###########################################
    elif selected_complaint == "Rt Hand (M79.641)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 7)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 9)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### L_Hand (M79.642) ########
###########################################
    elif selected_complaint == "Lt Hand(M79.642)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 7)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 9)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 9)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### Low back ##################
###########################################

    elif selected_complaint == "Low back (M54.50)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 13)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 12)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 7)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 9)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### R_Hip #####################
###########################################

    elif selected_complaint == "Rt Hip (M25.551)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 10)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 7)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### L_Hip #####################
###########################################

    elif selected_complaint == "Lt Hip (M25.552)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 10)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 10)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 7)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### R_K #######################
###########################################

    elif selected_complaint == "Rt Knee (M25.561)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 10)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### L_K #######################
###########################################

    elif selected_complaint == "Lt Knee (M25.562)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 10)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### R_A #######################
###########################################

    elif selected_complaint == "Rt Ankle (M25.571)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 4)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
############### L_A #######################
###########################################

    elif selected_complaint == "Lt Ankle (M25.572)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
########### R_Foot (M79.671) ##############
###########################################

    elif selected_complaint == "Rt Foot (M79.671)":
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

###########################################
########### L_Foot (M79.672) ##############
###########################################

    elif selected_complaint == "Lt Foot (M79.672)": 
        if selected_cpt == "G0283 Electrical Stimulation Unattended":
            procedure_text = df.iloc[1, 7]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            return  # Exit the function after processing

        elif selected_cpt == "97032 Electrical Stimulation Attended":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            procedure_text = df.iloc[1, 8]  # Row 3 (index 2) and Column H (index 7)
            for _ in range(int(selected_units)):  # Loop based on the number of units
                procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
            return  # Exit the function after processing

        elif selected_cpt == "97110 Therapeutic Exercise":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 9] for i in range(1, 11)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97112 Neuromuscular re-education":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 10] for i in range(1, 9)], 2)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97140 Manual Therapy":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 11] for i in range(1, 6)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97530 Therapeutic Activity":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 12] for i in range(1, 6)], 3)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

        elif selected_cpt == "97535 Self-Management":
            time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
            cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
            cpt_var.set("")  # Clear the CPT dropdown
            units_var.set("")  # Clear the Units dropdown
            units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
            random_cells = random.sample([df.iloc[i, 13] for i in range(1, 5)], 1)
            for cell in random_cells:
                for _ in range(int(selected_units)):  # Loop based on the number of units
                    procedure_table.insert("", "end", values=(selected_cpt, cell))
            return  # Exit the function after processing

################################
################################
    # This section goes last in the final loop after all conditions has not been met
    # Default logic for other CPT codes
    if selected_cpt == "97161 Initial Evaluation" and selected_units:
        time_minutes = int(selected_units) * 20 # 1 unit = 20 minutes
        cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
        cpt_var.set("")  # Clear the CPT dropdown
        units_var.set("")  # Clear the Units dropdown
        units_dropdown.current(0) # Resets the Units to display 1 instead of blank
        return
    if selected_cpt == "97164 Re-evaluation" and selected_units:
        time_minutes = int(selected_units) * 20 # 1 unit = 20 minutes
        cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
        cpt_var.set("")  # Clear the CPT dropdown
        units_var.set("")  # Clear the Units dropdown
        units_dropdown.current(0) # Resets the Units to display 1 instead of blank
        return
    elif selected_cpt and selected_units:
        time_minutes = int(selected_units) * 15  # 1 unit = 15 minutes
        cpt_tree.insert("", "end", values=(selected_cpt, selected_units, f"{time_minutes} mins"))
        cpt_var.set("")  # Clear the CPT dropdown
        units_var.set("")  # Clear the Units dropdown
        units_dropdown.current(0)  # Resets the Units to display 1 instead of blank
        return  # Exit the function after processing

def add_favorite_cpt(chief_complaint):
    """
    Adds a predefined set of favorite CPT codes based on the selected chief complaint.
    """
    # Clear the CPT tree and procedure table
    clear_all_cpt()
    clear_procedure()

    # Define the favorite CPT codes
    favorite_cpt_codes = [
        "97110 Therapeutic Exercise",
        "97112 Neuromuscular re-education",
        "97140 Manual Therapy",
        "G0283 Electrical Stimulation Unattended"
    ]

    # Add each CPT code to the CPT tree and procedure table
    for cpt_code in favorite_cpt_codes:
        # Set the CPT code and units in the dropdowns
        cpt_var.set(cpt_code)
        units_var.set("1")  # Default to 1 unit for each CPT code
        units_dropdown.current(0)  # Reset the units dropdown

        # Call the add_cpt_row function to add the CPT code
        add_cpt_row()

def add_favorite2_cpt(chief_complaint):
    """
    Adds a predefined set of favorite CPT codes based on the selected chief complaint.
    This function randomly selects 9 non-empty entries from the specified column and adds them to the procedure table.
    It reloads the data from the Excel sheet based on the current chief complaint selection.
    """
    # Clear the CPT tree and procedure table
    clear_all_cpt()
    clear_procedure()

    # Reload the data from the Excel sheet based on the selected chief complaint
    selected_complaint = chief_complaint_var.get()
    if current_file_path:
        try:
            global df
            df = pd.read_excel(current_file_path, sheet_name=selected_complaint)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet '{selected_complaint}': {e}")
            return

    # Define the CPT code and units
    selected_cpt = "97530 Therapeutic Activity"
    units = "3"

    # Add the CPT code with 3 units to the CPT table
    time_minutes = int(units) * 15  # 1 unit = 15 minutes
    cpt_tree.insert("", "end", values=(selected_cpt, units, f"{time_minutes} mins"))

    # Get the column index for the procedure text (column 12, index 11)
    column_index = 12

    # Collect all non-empty entries from the specified column
    non_empty_entries = []
    for i in range(1, len(df)):  # Start from row 1 to skip the header
        procedure_text = df.iloc[i, column_index]
        if pd.notna(procedure_text) and procedure_text.strip():  # Check if the cell is not empty
            non_empty_entries.append(procedure_text)

    # Randomly select 9 entries (or fewer if there are not enough entries)
    num_entries_to_select = min(3, len(non_empty_entries))  # Ensure we don't exceed the available entries
    if num_entries_to_select > 0:
        random_entries = random.sample(non_empty_entries, num_entries_to_select)
    # else:
    #     messagebox.showwarning("Warning", "No non-empty entries found in the specified column.")
    #     return

    # Insert the randomly selected entries into the procedure table
    for i, procedure_text in enumerate(random_entries):
        if i == 0:
            # Insert the CPT code only for the first row
            procedure_table.insert("", "end", values=(selected_cpt, procedure_text))
        else:
            # Insert empty CPT code for subsequent rows
            procedure_table.insert("", "end", values=("", procedure_text))

def remove_cpt_row():
    # Get the selected item from the CPT tree
    selected_item = cpt_tree.selection()
    
    if selected_item:
        # Get the CPT code from the selected row in the CPT tree
        selected_cpt = cpt_tree.item(selected_item, "values")[0]
        
        # Remove the selected row from the CPT tree
        cpt_tree.delete(selected_item)
        
        for row in procedure_table.get_children():
            values = procedure_table.item(row, "values")
            if values[0] == selected_cpt:  # Check if the CPT code matches
                procedure_table.delete(row)
# Function to clear all entries in the CPT code tree
def clear_all_cpt():
    for row in cpt_tree.get_children():
        cpt_tree.delete(row)

######################
# UI Hover
######################
# Function to delete a row from the CPT code table
def delete_cpt_row():
    # Trigger the existing "Remove CPT" function
    remove_cpt_row()

# Function to edit a row in the CPT code table
def edit_cpt_row():
    selected = cpt_tree.selection()
    if selected:
        # Get the current values of the selected row
        values = cpt_tree.item(selected, "values")
        
        # Open a dialog to edit the row
        new_name = simpledialog.askstring("Edit", "Enter new CPT code name:", initialvalue=values[0])
        if new_name:
            # Update the row with the new name
            cpt_tree.item(selected, values=(new_name, values[1], values[2]))

# Function to show the context menu on hover for CPT code table
def show_cpt_context_menu(event):
    # Identify the row under the mouse pointer
    row_id = cpt_tree.identify_row(event.y)
    if row_id:
        # Select the row
        cpt_tree.selection_set(row_id)
        
        # Show the context menu at the mouse position
        cpt_context_menu.post(event.x_root, event.y_root)

# Function to hide the context menu for CPT code table
def hide_cpt_context_menu(event):
    cpt_context_menu.unpost()

#############################################

# Function to delete a row from the procedure table
def delete_procedure_row():
    selected = procedure_table.selection()
    if selected:
        procedure_table.delete(selected)

# Function to edit a row in the procedure table
def edit_procedure_row():
    selected = procedure_table.selection()
    if selected:
        # Get the current values of the selected row
        values = procedure_table.item(selected, "values")
        
        # Open a dialog to edit the row
        new_name = simpledialog.askstring("Edit", "Enter new procedure name:", initialvalue=values[1])
        if new_name:
            # Update the row with the new name
            procedure_table.item(selected, values=(values[0], new_name))

# Function to show the context menu on hover
def show_context_menu(event):
    # Identify the row under the mouse pointer
    row_id = procedure_table.identify_row(event.y)
    if row_id:
        # Select the row
        procedure_table.selection_set(row_id)
        
        # Show the context menu at the mouse position
        context_menu.post(event.x_root, event.y_root)

# Function to hide the context menu
def hide_context_menu(event):
    context_menu.unpost()

def reset_all_fields():
    """
    Resets all fields to their default state, allowing the user to start a new progress note.
    """
    # Clear patient information
    patient_entry.delete(0, tk.END)
    dob_entry.delete(0, tk.END)
    # insurance_dropdown.current(0)
    insurance_var.set("")  # Set Insurance to empty
    chief_complaint_var.set("")  # Set Chief Complaint to empty

    # Reset chief complaint dropdown
    # chief_complaint_dropdown.current(0 if chief_complaint_dropdown['values'] else -1)

    # Clear CPT and units dropdowns
    cpt_var.set("")
    units_var.set("")
    units_dropdown.current(0)
    # company_dropdown.current(0)

    # Clear CPT tree and procedure table
    for row in cpt_tree.get_children():
        cpt_tree.delete(row)
    for row in procedure_table.get_children():
        procedure_table.delete(row)

    # Clear Subjective and Objective text fields
    subjective_text.delete(1.0, tk.END)
    objective_text.delete(1.0, tk.END)

    # Reset Assessment dropdown
    assessment_dropdown.current(1)

    # Clear Plan text field and reset to default text
    plan_text.delete(1.0, tk.END)
    plan_text.insert(tk.END, "It is recommended that the patient continue physical therapy targeting the affected areas at a frequency of 2-3 sessions per week, with a follow-up re-evaluation in 4-5 weeks to assess progress and guide ongoing therapeutic management.")

    # Reset therapist dropdowns
    # therapist_dropdown_top.current(0)
    # therapist_dropdown_bottom.current(0)

    # Pre-fill DOS and therapist's date with today's date
    prefill_dates()

    # Place cursor in the Patient text box
    patient_entry.focus_set()

def reset_fields():
    """
    Clears the CPT tree table, procedure tree table, subjective text field, and objective text field.
    """
    # Clear the CPT tree table
    for row in cpt_tree.get_children():
        cpt_tree.delete(row)

    # Clear the procedure tree table
    for row in procedure_table.get_children():
        procedure_table.delete(row)

    # Clear the subjective text field
    subjective_text.delete(1.0, tk.END)

    # Clear the objective text field
    objective_text.delete(1.0, tk.END)

# duplicated?
# def load_sheet_based_on_chief_complaint():
#     """
#     Loads the sheet based on the selected chief complaint and resets the fields.
#     """
#     global sheet
#     selected_complaint = chief_complaint_var.get()
#     if selected_complaint:
#         try:
#             sheet = workbook[selected_complaint]  # Dynamically select the sheet based on the chief complaint
#         except Exception as e:
#             messagebox.showerror("Error", f"Failed to load sheet '{selected_complaint}': {e}")

##################################
# End of UI Hover
##################################

# Function to save the progress note directly to a PDF file
def save_to_pdf(patient_entry, dob_entry, chief_complaint_var, cpt_tree, subjective_text, objective_text, procedure_table, assessment_var, plan_text, therapist_var, therapist_date_entry):
    """
    Saves the progress note to a PDF file after validating for empty fields, using the selected address from address_var.
    """
    # List of required fields and their corresponding widgets
    required_fields = [
        ("Patient Name", patient_entry),
        ("Date of Service", dos_entry),
        ("Date of Birth", dob_entry),
        ("Chief Complaint", chief_complaint_var),
        ("Subjective", subjective_text),
        ("Objective", objective_text),
        ("Assessment", assessment_var),
        ("Plan", plan_text),
        ("Therapist Name", therapist_var),
        ("Therapist Date", therapist_date_entry),
    ]

    # Check for empty fields
    empty_fields = []
    for field_name, widget in required_fields:
        if isinstance(widget, tk.Entry):
            if not widget.get().strip():
                empty_fields.append(field_name)
        elif isinstance(widget, tk.Text):
            if not widget.get("1.0", tk.END).strip():
                empty_fields.append(field_name)
        elif isinstance(widget, tk.StringVar):
            if not widget.get().strip():
                empty_fields.append(field_name)

    # If there are empty fields, notify the user and ask if they want to continue
    if empty_fields:
        empty_fields_str = ", ".join(empty_fields)
        response = messagebox.askyesno(
            "Empty Fields",
            f"The following fields are empty: {empty_fields_str}. Do you want to continue?",
        )
        if not response:
            return  # User chose not to continue

    # Proceed with saving the PDF
    dos_date_str = dos_entry.get()  # Get the date string (e.g., "03/31/2025")
    try:
        dos_date_obj = datetime.strptime(dos_date_str, "%m/%d/%Y")  # Parse "MM/DD/YYYY"
        dos_date = dos_date_obj.strftime("%Y%m%d")  # Convert to "YYYYMMDD" (e.g., "20250331")
    except ValueError:
        dos_date = dos_date_str.replace("/", "")  # Fallback to original behavior if parsing fails

    name_only = therapist_var.get().split(",")[0].strip()
    insurance_name = insurance_var.get()
    patient_name = patient_entry.get()
    default_file_name = f"{patient_name} ({insurance_name}) {dos_date}.pdf"

    # Open the save file dialog with the default file name
    file_path = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF Document", "*.pdf")],
        initialdir="~/Documents",
        title="Save Progress Note As PDF",
        initialfile=default_file_name
    )
    if not file_path:
        return  # User canceled the save dialog

    # Create a PDF document with reduced margins
    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter,
        leftMargin=29,
        rightMargin=24,
        topMargin=25,
        bottomMargin=40
    )
    styles = getSampleStyleSheet()

    # Custom style definitions (unchanged)
    normal_style = ParagraphStyle("Normal", parent=styles["Normal"], fontSize=10, leading=12, spaceAfter=3)
    bold_style = ParagraphStyle("Bold", parent=styles["Normal"], fontSize=10, leading=12, spaceAfter=6, fontName="Helvetica-Bold")
    indented_style = ParagraphStyle("Indented", parent=styles["Normal"], fontSize=10, leading=12, spaceAfter=3, leftIndent=10)
    centered_style = ParagraphStyle("Centered", parent=styles["Normal"], fontSize=10, leading=12, spaceAfter=6, alignment=1)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=12, leading=14, spaceAfter=2)
    header_style = ParagraphStyle("Header", parent=styles["Title"], fontSize=14, leading=16, alignment=1, spaceAfter=1)
    contact_style = ParagraphStyle("Contact", parent=styles["Normal"], fontSize=8, leading=10, alignment=1, spaceAfter=12, fontName="Helvetica-Bold")

    # Content list to hold all elements
    content = []

    # Add the header with dynamic therapist name
    therapist_name = therapist_var.get()
    company_name = company_var.get()
    header_text = f"{company_name}"
    content.append(Paragraph(header_text, header_style))
    content.append(Spacer(1, 6))

    # Contact info line using the selected address from address_var
    selected_address = address_var.get()  # Get the address from the dropdown
    company_info = company_info_dict.get(selected_address, {"address": "", "telephone": "", "fax": ""})
    contact_info = f"{selected_address} \u00A0 \u00A0 Tel: {company_info['telephone']} \u00A0 \u00A0 Fax: {company_info['fax']}"
    content.append(Paragraph(contact_info, contact_style))

    # Rest of the PDF generation code remains unchanged
    patient_info_data = [
        [Paragraph(f"<b>Patient:</b> {patient_entry.get()}", normal_style), 
         Paragraph(f"<b>DOS:</b> {dos_entry.get()}", normal_style)],
        [Paragraph(f"<b>DOB:</b> {dob_entry.get()}", normal_style), 
         Paragraph(f"<b>Ins:</b> {insurance_name}", normal_style)],
        [Paragraph(f"<b>Chief Complaint:</b> {chief_complaint_var.get()}", normal_style)]
    ]

    patient_info_table = Table(patient_info_data, colWidths=[380, 180])
    patient_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
    ]))
    content.append(patient_info_table)
    content.append(Spacer(1, 2))

    combined_data = [[""], ["Billing Code"]]
    combined_table = Table(combined_data, colWidths=[doc.width])
    combined_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.gray),
        ('ALIGN', (0, 1), (-1, 1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 2),
        ('TOPPADDING', (0, 1), (-1, 1), 2),
    ]))
    content.append(combined_table)

    data = [["CPT Code", "Units", "Time"]]
    for child in cpt_tree.get_children():
        values = cpt_tree.item(child, 'values')
        data.append([values[0], values[1], values[2]])

    table = Table(data, colWidths=[200, 40, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    content.append(table)
    content.append(Spacer(1, 2))

    content.append(Paragraph("Subjective", bold_style))
    subjective_content = subjective_text.get("1.0", tk.END).strip()
    for line in subjective_content.split("\n"):
        if line.strip():
            content.append(Paragraph(line.strip(), indented_style))
    content.append(Spacer(1, 6))

    content.append(Paragraph("Objective", bold_style))
    objective_content = objective_text.get("1.0", tk.END).strip()
    for line in objective_content.split("\n"):
        if line.strip():
            content.append(Paragraph(line.strip(), indented_style))
    content.append(Spacer(1, 6))

    content.append(Paragraph("Procedure", bold_style))
    cpt_code_dict = {}
    for child in procedure_table.get_children():
        values = procedure_table.item(child, 'values')
        cpt_code = values[0]
        procedure_text = values[1]
        if cpt_code not in cpt_code_dict:
            cpt_code_dict[cpt_code] = []
        cpt_code_dict[cpt_code].append(procedure_text)

    for cpt_code, procedures in cpt_code_dict.items():
        time_minutes = ""
        for child in cpt_tree.get_children():
            values = cpt_tree.item(child, 'values')
            if values[0] == cpt_code:
                time_minutes = values[2]
                break
        # content.append(Paragraph(f"<b>{cpt_code}</b> <b>{time_minutes}</b>", indented_style))
        # for procedure in procedures:
        #     content.append(Paragraph(f"&nbsp;&nbsp;{procedure}", indented_style))
        # content.append(Spacer(1, 6))
# Add the CPT code and time as a separate bolded line

        content.append(Paragraph(f"<b>{cpt_code}</b> <b>{time_minutes}</b>", indented_style))
        
        # Combine all procedures under this CPT code into a single paragraph with consistent indentation
        procedure_text_combined = "<br/>".join([f"&nbsp;&nbsp;{proc}" for proc in procedures if proc])
        content.append(Paragraph(procedure_text_combined, indented_style))
        
        content.append(Spacer(1, 6))

    assessment_content = f"<b>Assessment:</b> \u00A0 \u00A0 {assessment_var.get()} Progress Made"
    content.append(Paragraph(assessment_content, normal_style))
    content.append(Spacer(1, 6))

    content.append(Paragraph("Plan", bold_style))
    plan_content = plan_text.get("1.0", tk.END).strip()
    content.append(Paragraph(f"{plan_content}", indented_style))
    content.append(Spacer(1, 12))

    therapist_name = therapist_var.get()
    therapist_license_number = therapist_license_dict.get(therapist_name, "N/A")
    int_therapist_license_number = therapist_license_number.zfill(6)
    evaluated_by_text = f"This patient was evaluated and treated by {therapist_name} #{int_therapist_license_number} on {therapist_date_entry.get()}"
    content.append(Paragraph(evaluated_by_text, normal_style))
    content.append(Spacer(1, 6))

    # def footer(canvas, doc):
    #     canvas.saveState()
    #     signature_text = f"E-signed by {therapist_var.get()} on {therapist_date_entry.get()}"
    #     canvas.setFont("Helvetica", 10)
    #     canvas.drawCentredString(doc.width / 2.0, 20, signature_text)
    #     canvas.restoreState()

    # doc.build(content, onFirstPage=footer, onLaterPages=footer)
    # messagebox.showinfo("Success", "Progress note saved to PDF file.")


    def footer(canvas, doc):
        canvas.saveState()
        signature_text = f"E-signed by {esigned_var.get()} on {esigned_date_entry.get()}"
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(doc.width / 2.0, 20, signature_text)
        canvas.restoreState()

    doc.build(content, onFirstPage=footer, onLaterPages=footer)
    messagebox.showinfo("Success", "Progress note saved to PDF file.")

# Main application window
root = tk.Tk()
root.title("Progress Note Generator")

# Set the custom icon (place this right after root is defined)
root.iconbitmap(resource_path("custom_icon.ico"))  # For .ico files

# Define therapist_var globally before UI setup
therapist_var = tk.StringVar()  # Added here to fix the NameError

# Add this after therapist_var definition
esigned_var = tk.StringVar()

# Function to update the E-signed by dropdown when Company changes
def update_esigned_dropdown(*args):
    selected_company = company_var.get()
    esigners = esigners_by_company.get(selected_company, [])
    esigned_dropdown['values'] = esigners
    if esigners:
        esigned_var.set(esigners[0])
    else:
        esigned_var.set("")

# Function to pre-fill DOS and therapist's date with today's date
def prefill_dates():
    today = datetime.now().strftime("%m/%d/%Y")
    dos_entry.delete(0, tk.END)
    dos_entry.insert(0, today)
    therapist_date_entry.delete(0, tk.END)
    therapist_date_entry.insert(0, today)
    esigned_date_entry.delete(0, tk.END)  # Add this line
    esigned_date_entry.insert(0, today)   # Add this line

# Define a smaller font for compactness
modern_font = ("Helvetica", 10)  # Reduced font size

# Top Section: Therapist, Company, Address
# tk.Label(root, text="Therapist:", font=("Helvetica", 9)).grid(row=3, column=0, sticky="e", padx=2, pady=2)
# therapist_var = tk.StringVar()
# therapist_dropdown_top = ttk.Combobox(root, textvariable=therapist_var, width=20, font=("Helvetica", 9))
# therapist_dropdown_top.grid(row=3, column=1, padx=2, pady=2, sticky="w")

tk.Label(root, text="Company:", font=("Helvetica", 9)).grid(row=0, column=0, sticky="e", padx=2, pady=2)
company_var = tk.StringVar()
company_dropdown = ttk.Combobox(root, textvariable=company_var, width=28, font=("Helvetica", 9))
company_dropdown.grid(row=0, column=1, padx=2, pady=2, sticky="w")

tk.Label(root, text="Address:", font=("Helvetica", 9)).grid(row=0, column=2, sticky="e", padx=2, pady=2)
address_var = tk.StringVar()
address_dropdown = ttk.Combobox(root, textvariable=address_var, width=29, font=("Helvetica", 9))
address_dropdown.grid(row=0, column=3, padx=2, pady=2, sticky="w")

# Patient Information Section
# tk.Label(root, text="Patient:", font=("Helvetica", 9)).grid(row=1, column=0, sticky="e", padx=2, pady=2)
# patient_entry = tk.Entry(root, width=28, font=("Helvetica", 9))
# patient_entry.grid(row=1, column=1, padx=2, pady=2, sticky="w")

# Patient Information Section
tk.Label(root, text="Patient:", font=("Helvetica", 9)).grid(row=1, column=0, sticky="e", padx=2, pady=2)
patient_entry = tk.Entry(root, width=28, font=("Helvetica", 9))
patient_entry.grid(row=1, column=1, padx=2, pady=2, sticky="w", ipady=2)  # Added ipady=2 for extra vertical space

tk.Label(root, text="DOS:", font=("Helvetica", 9)).grid(row=1, column=2, sticky="e", padx=2, pady=2)
dos_entry = ttk.Entry(root, width=15, font=("Helvetica", 9))
dos_entry.grid(row=1, column=3, padx=2, pady=2, sticky="w")
ttk.Button(root, text="Pick Date", command=open_dos_calendar).grid(row=1, column=4, padx=2, pady=2)

tk.Label(root, text="DOB:", font=("Helvetica", 9)).grid(row=2, column=2, sticky="e", padx=2, pady=2)
dob_entry = ttk.Entry(root, width=15, font=("Helvetica", 9))
dob_entry.grid(row=2, column=3, padx=2, pady=2, sticky="w")
ttk.Button(root, text="Pick Date", command=open_calendar).grid(row=2, column=4, padx=2, pady=2)

tk.Label(root, text="Insurance:", font=("Helvetica", 9)).grid(row=2, column=0, sticky="e", padx=2, pady=2)
insurance_var = tk.StringVar()
insurance_dropdown = ttk.Combobox(root, textvariable=insurance_var, width=20, font=("Helvetica", 9))
insurance_dropdown.grid(row=2, column=1, padx=2, pady=2, sticky="w")

# Chief Complaint Section
tk.Label(root, text="Chief Complaint:", font=("Helvetica", 9)).grid(row=4, column=0, sticky="e", padx=2, pady=2)
chief_complaint_var = tk.StringVar()
chief_complaint_dropdown = ttk.Combobox(root, textvariable=chief_complaint_var, width=20, font=("Helvetica", 9))
chief_complaint_dropdown.grid(row=4, column=1, columnspan=1, padx=2, pady=2, sticky="w")
# Remove chief_complaint_var.trace_add("write", lambda *args: reset_fields()) if you don’t want fields reset on every change
chief_complaint_var.trace_add("write", lambda *args: load_sheet_based_on_chief_complaint())


# CPT Code and Units Section
tk.Label(root, text="CPT Code:", font=("Helvetica", 9)).grid(row=5, column=0, sticky="e", padx=2, pady=2)
cpt_var = tk.StringVar()
cpt_dropdown = ttk.Combobox(root, textvariable=cpt_var, values=cpt_codes, width=28, font=("Helvetica", 9))
cpt_dropdown.grid(row=5, column=1, padx=2, pady=2, sticky="w")
cpt_dropdown.current(0)

# tk.Label(root, text="Units:", font=("Helvetica", 9)).grid(row=5, column=2, sticky="e", padx=2, pady=2)
# units_var = tk.StringVar()
# units_dropdown = ttk.Combobox(root, textvariable=units_var, values=units_options, width=5, font=("Helvetica", 9))
# units_dropdown.grid(row=5, column=3, padx=2, pady=2)
# units_dropdown.current(0)

tk.Label(root, text="Units:", font=("Helvetica", 9)).grid(row=5, column=2, sticky="e", padx=(2, 0), pady=2)  # Reduced padx on right
units_var = tk.StringVar()
units_dropdown = ttk.Combobox(root, textvariable=units_var, values=units_options, width=5, font=("Helvetica", 9))
units_dropdown.grid(row=5, column=3, padx=(0, 2), pady=2, sticky="w")  # Reduced padx on left, aligned left
units_dropdown.current(0)

cpt_frame = ttk.Frame(root)
cpt_frame.grid(row=6, column=1, columnspan=3, padx=2, pady=2, sticky="ew")
columns = ("CPT Code", "Units", "Time")
cpt_tree = ttk.Treeview(cpt_frame, columns=columns, show="headings", height=4)
cpt_tree.heading("CPT Code", text="CPT Code")
cpt_tree.heading("Units", text="Units")
cpt_tree.heading("Time", text="Time")
cpt_tree.column("CPT Code", width=120)
cpt_tree.column("Units", width=40, anchor="center")
cpt_tree.column("Time", width=60, anchor="center")
cpt_tree.pack(side="left", fill="both", expand=True)

cpt_scroll = ttk.Scrollbar(cpt_frame, orient="vertical", command=cpt_tree.yview)
cpt_tree.configure(yscrollcommand=cpt_scroll.set)
cpt_scroll.pack(side="right", fill="y")

# CPT Buttons (including Favorite #1 and #2)
add_cpt_button = ttk.Button(root, text="Add", command=add_cpt_row)
add_cpt_button.grid(row=5, column=4, padx=2, pady=2)

favorite_button = tk.Button(root, text="Fav #1", command=lambda: add_favorite_cpt(chief_complaint_var.get()), font=("Helvetica", 9))
favorite_button.grid(row=5, column=5, padx=2, pady=2)

favorite_button_2 = tk.Button(root, text="Fav #2", command=lambda: add_favorite2_cpt(chief_complaint_var.get()), font=("Helvetica", 9))
favorite_button_2.grid(row=5, column=6, padx=2, pady=2)

remove_cpt_button = ttk.Button(root, text="Remove", command=remove_cpt_row)
remove_cpt_button.grid(row=6, column=4, padx=2, pady=2)

clear_all_cpt_button = ttk.Button(root, text="Clear", command=clear_all_cpt)
clear_all_cpt_button.grid(row=6, column=5, padx=2, pady=2)

# Severity Radio Buttons
tk.Label(root, text="Severity:", font=("Helvetica", 9)).grid(row=7, column=0, sticky="e", padx=2, pady=2)
severity_var = tk.StringVar(value="9-10")
severity_frame = ttk.Frame(root)
severity_frame.grid(row=7, column=1, columnspan=2, padx=2, pady=2, sticky="w")
for level in ["9-10", "7-8", "5-6", "4-5", "2-3"]:
    tk.Radiobutton(severity_frame, text=level, variable=severity_var, value=level, font=("Helvetica", 9)).pack(side="left", padx=8)

# Subjective and Objective Sections
tk.Label(root, text="Subjective:", font=("Helvetica", 9)).grid(row=8, column=0, sticky="e", padx=2, pady=2)
subjective_frame = ttk.Frame(root)
subjective_frame.grid(row=8, column=1, columnspan=3, padx=2, pady=2, sticky="ew")
subjective_text = tk.Text(subjective_frame, height=4, width=40, wrap=tk.WORD, font=modern_font)
subjective_text.pack(side="left", fill="both", expand=True)
subjective_scroll = ttk.Scrollbar(subjective_frame, orient="vertical", command=subjective_text.yview)
subjective_text.configure(yscrollcommand=subjective_scroll.set)
subjective_scroll.pack(side="right", fill="y")

tk.Label(root, text="Objective:", font=("Helvetica", 9)).grid(row=9, column=0, sticky="e", padx=2, pady=2)
objective_frame = ttk.Frame(root)
objective_frame.grid(row=9, column=1, columnspan=3, padx=2, pady=2, sticky="ew")
objective_text = tk.Text(objective_frame, height=4, width=40, wrap=tk.WORD, font=modern_font)
objective_text.pack(side="left", fill="both", expand=True)
objective_scroll = ttk.Scrollbar(objective_frame, orient="vertical", command=objective_text.yview)
objective_text.configure(yscrollcommand=objective_scroll.set)
objective_scroll.pack(side="right", fill="y")

# Procedure Section
tk.Label(root, text="Procedure:", font=("Helvetica", 9)).grid(row=10, column=0, sticky="e", padx=2, pady=2)
procedure_frame = ttk.Frame(root)
procedure_frame.grid(row=10, column=1, columnspan=3, padx=2, pady=2, sticky="ew")
procedure_table = ttk.Treeview(procedure_frame, columns=("CPT Code", "Procedure"), show="headings", height=7)
procedure_table.heading("CPT Code", text="CPT Code")
procedure_table.heading("Procedure", text="Procedure")
procedure_table.column("CPT Code", width=100)
procedure_table.column("Procedure", width=250)
procedure_table.pack(side="left", fill="both", expand=True)
procedure_scroll = ttk.Scrollbar(procedure_frame, orient="vertical", command=procedure_table.yview)
procedure_table.configure(yscrollcommand=procedure_scroll.set)
procedure_scroll.pack(side="right", fill="y")

# Context Menu for CPT Tree
cpt_context_menu = Menu(root, tearoff=0)
cpt_context_menu.add_command(label="Delete", command=delete_cpt_row)
cpt_context_menu.add_command(label="Edit", command=edit_cpt_row)

# Context Menu for Procedure Table
context_menu = Menu(root, tearoff=0)
context_menu.add_command(label="Delete", command=delete_procedure_row)
context_menu.add_command(label="Edit", command=edit_procedure_row)

# Assessment and Plan
tk.Label(root, text="Assessment:", font=("Helvetica", 9)).grid(row=11, column=0, sticky="e", padx=2, pady=2)
assessment_var = tk.StringVar()
assessment_dropdown = ttk.Combobox(root, textvariable=assessment_var, values=assessment_options, width=7, font=("Helvetica", 9))
assessment_dropdown.grid(row=11, column=1, padx=2, pady=2)
assessment_dropdown.current(2)
tk.Label(root, text="Progress Made", font=("Helvetica", 9)).grid(row=11, column=2, sticky="w", padx=2, pady=2)

tk.Label(root, text="Plan:", font=("Helvetica", 9)).grid(row=12, column=0, sticky="e", padx=2, pady=2)
plan_frame = ttk.Frame(root)
plan_frame.grid(row=12, column=1, columnspan=3, padx=2, pady=2, sticky="ew")
plan_text = tk.Text(plan_frame, height=4, width=40, wrap=tk.WORD, font=modern_font)
plan_text.pack(side="left", fill="both", expand=True)
plan_scroll = ttk.Scrollbar(plan_frame, orient="vertical", command=plan_text.yview)
plan_text.configure(yscrollcommand=plan_scroll.set)
plan_scroll.pack(side="right", fill="y")
plan_text.insert(tk.END, "It is recommended that the patient continue physical therapy targeting the affected areas at a frequency of 2-3 sessions per week, with a follow-up re-evaluation in 4-5 weeks to assess progress and guide ongoing therapeutic management.")

# Therapist Bottom Section
tk.Label(root, text="Evaluating Therapist:", font=("Helvetica", 9)).grid(row=13, column=0, sticky="e", padx=2, pady=2)
therapist_dropdown_bottom = ttk.Combobox(root, textvariable=therapist_var, width=20, font=("Helvetica", 9))
therapist_dropdown_bottom.grid(row=13, column=1, padx=2, pady=2)
therapist_date_entry = ttk.Entry(root, width=12, font=("Helvetica", 9))
therapist_date_entry.grid(row=13, column=2, padx=2, pady=2)
ttk.Button(root, text="Pick Date", command=open_therapist_calendar).grid(row=13, column=3, padx=2, pady=2)

# New "E-signed by:" section with date entry and button
tk.Label(root, text="E-signed by:", font=("Helvetica", 9)).grid(row=14, column=0, sticky="e", padx=2, pady=2)
esigned_dropdown = ttk.Combobox(root, textvariable=esigned_var, width=20, font=("Helvetica", 9))
esigned_dropdown.grid(row=14, column=1, padx=2, pady=2)
esigned_date_entry = ttk.Entry(root, width=12, font=("Helvetica", 9))
esigned_date_entry.grid(row=14, column=2, padx=2, pady=2)
ttk.Button(root, text="Pick Date", command=open_esigned_calendar).grid(row=14, column=3, padx=2, pady=2)

# Link the Company dropdown to update E-signed by dropdown
company_var.trace_add("write", update_esigned_dropdown)

# Action Buttons
generate_so_button = tk.Button(root, text="Generate SO", command=display_scenario, font=("Helvetica", 9))
generate_so_button.grid(row=7, column=4, padx=2, pady=2)

clear_scenarios_button = tk.Button(root, text="Clear SO", command=clear_scenarios, font=("Helvetica", 9))
clear_scenarios_button.grid(row=7, column=5, padx=2, pady=2)

clear_procedure_button = tk.Button(root, text="Clear Procedure", command=clear_procedure, font=("Helvetica", 9))
clear_procedure_button.grid(row=10, column=4, padx=2, pady=2)

save_pdf_button = tk.Button(root, text="Save to PDF", command=lambda: save_to_pdf(patient_entry, dob_entry, chief_complaint_var, cpt_tree, subjective_text, objective_text, procedure_table, assessment_var, plan_text, therapist_var, therapist_date_entry), font=("Helvetica", 9))
save_pdf_button.grid(row=15, column=2, padx=2, pady=5)

reset_all_button = tk.Button(root, text="New Patient", command=reset_all_fields, font=("Helvetica", 9))
reset_all_button.grid(row=15, column=3, padx=2, pady=5)

return_to_login_button = tk.Button(root, text="Logout", command=return_to_login, font=("Helvetica", 9))
return_to_login_button.grid(row=15, column=1, padx=2, pady=5)

exit_button = tk.Button(root, text="Exit", command=exit_program, font=("Helvetica", 9))
exit_button.grid(row=15, column=4, padx=2, pady=5)

# Add a small empty row for additional padding below the buttons
tk.Label(root, text="", height=0).grid(row=16, column=0, pady=1)  # Empty label with small padding

# Bind events
cpt_tree.bind("<Button-3>", show_cpt_context_menu)
cpt_tree.bind("<ButtonPress-1>", start_drag)
cpt_tree.bind("<B1-Motion>", on_drag)
cpt_tree.bind("<ButtonRelease-1>", stop_drag)
procedure_table.bind("<Button-3>", show_context_menu)
procedure_table.bind("<ButtonPress-1>", start_drag_procedure)
procedure_table.bind("<B1-Motion>", on_drag_procedure)
procedure_table.bind("<ButtonRelease-1>", stop_drag_procedure)

# Pre-fill dates and start login screen
prefill_dates()
login_screen()

# Start the main event loop
root.mainloop()
